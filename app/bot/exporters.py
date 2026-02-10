from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Dict, List, Tuple
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


def build_transactions_xlsx(
    transactions: List[Dict[str, object]],
    tz_name: str = "America/Bogota",
) -> Tuple[bytes, str]:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Transacciones"

    headers = [
        "Fecha",
        "Tipo",
        "Clase",
        "Monto",
        "Moneda",
        "Categoria",
        "Descripcion",
        "Comercio",
        "Metodo",
        "Contraparte",
        "Rol prestamo",
        "Recurrente",
        "Recurrencia",
        "Creado",
        "Actualizado",
        "TxId",
    ]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    rows = [tx for tx in transactions if not tx.get("isDeleted")]
    rows.sort(key=_tx_sort_key, reverse=True)

    for tx in rows:
        sheet.append(
            [
                _safe_str(tx.get("date")),
                _safe_str(tx.get("type")),
                _safe_str(tx.get("transactionKind")),
                _safe_float(tx.get("amount")),
                _safe_str(tx.get("currency")),
                _safe_str(tx.get("category")),
                _safe_str(tx.get("description")),
                _safe_str(tx.get("normalizedMerchant")),
                _safe_str(tx.get("paymentMethod")),
                _safe_str(tx.get("counterparty")),
                _safe_str(tx.get("loanRole")),
                "yes" if tx.get("isRecurring") else "no",
                _safe_str(tx.get("recurrence")),
                _safe_str(tx.get("createdAt")),
                _safe_str(tx.get("updatedAt")),
                _safe_str(tx.get("txId")),
            ]
        )

    sheet.freeze_panes = "A2"
    _autosize_columns(sheet, len(headers))

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    today = datetime.now(ZoneInfo(tz_name)).strftime("%Y%m%d")
    filename = f"transacciones_{today}.xlsx"
    return buffer.read(), filename


def _safe_str(value: object) -> str:
    if value is None:
        return ""
    return str(value)


def _safe_float(value: object) -> float:
    try:
        return float(str(value))
    except Exception:
        return 0.0


def _tx_sort_key(tx: Dict[str, object]) -> float:
    date_value = str(tx.get("date") or "")
    if date_value and len(date_value) == 10:
        try:
            return datetime.fromisoformat(date_value + "T00:00:00+00:00").timestamp()
        except ValueError:
            pass
    created_at = str(tx.get("createdAt") or "")
    try:
        return datetime.fromisoformat(created_at.replace("Z", "+00:00")).timestamp()
    except ValueError:
        return 0.0


def _autosize_columns(sheet, column_count: int) -> None:
    for col in range(1, column_count + 1):
        max_len = 0
        for cell in sheet[get_column_letter(col)]:
            value = cell.value
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))
        sheet.column_dimensions[get_column_letter(col)].width = min(max_len + 2, 40)
